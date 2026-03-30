"""
Claudette Bot - Entry Point Modular.
Todos los handlers: texto, voz, foto, ubicación, comandos, recordatorios.
"""

import os
import io
import re
import base64
import tempfile
import pytz
from datetime import datetime, timedelta

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes
)

from config import (
    TELEGRAM_BOT_TOKEN, OPENAI_API_KEY, ELEVENLABS_API_KEY,
    ELEVENLABS_VOICE_ID, OWNER_CHAT_ID, DEFAULT_LOCATION, logger
)
from brain import process_chat, conversation_history, user_modes, build_system_prompt, generate_morning_summary, generate_weekly_synthesis
from tools_registry import (
    user_locations, get_weather, search_web_google
)
from utils_security import restricted, get_youtube_transcript
from memory_manager import get_all_facts, save_fact, get_fact, setup_database

# --- Inicializar base de datos ---
setup_database()

# --- Inicializar tablas KB extra (document_links, mental_model_usage) ---
try:
    from knowledge_base import setup_kb_extra_tables
    setup_kb_extra_tables()
except Exception as _e:
    pass

# --- Clients opcionales ---
openai_client = None
elevenlabs_client = None

if OPENAI_API_KEY:
    from openai import OpenAI
    openai_client = OpenAI(api_key=OPENAI_API_KEY)

if ELEVENLABS_API_KEY:
    from elevenlabs.client import ElevenLabs
    elevenlabs_client = ElevenLabs(api_key=ELEVENLABS_API_KEY)

# Google services ya no necesarios aquí — brain.py los importa para el resumen matutino


# =====================================================
# MENÚ VISUAL
# =====================================================

async def show_menu(update, context):
    keyboard = [
        [
            InlineKeyboardButton("Boletin Matutino", callback_data='btn_morning'),
            InlineKeyboardButton("Noticias", callback_data='btn_news'),
        ],
        [
            InlineKeyboardButton("Progreso & Stats", callback_data='btn_progreso'),
            InlineKeyboardButton("Sintesis Semanal", callback_data='btn_sintesis'),
        ],
        [
            InlineKeyboardButton("Modo Profundo", callback_data='btn_deep'),
            InlineKeyboardButton("Modo Normal", callback_data='btn_normal'),
        ],
        [
            InlineKeyboardButton("Ver Memoria", callback_data='btn_mem'),
            InlineKeyboardButton("Borrar Chat", callback_data='btn_clear'),
        ],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    nl = chr(10)
    capacidades = (
        "CLAUDETTE v2 - Centro de Control" + nl + nl
        + "COMANDOS:" + nl
        + "/buenosdias  Boletin matutino (noticias + Midas + libro)" + nl
        + "/noticias    Noticias curadas en tiempo real" + nl
        + "/progreso    Panel: KB + biblioteca + modelos mentales" + nl
        + "/sintesis    Sintesis semanal de aprendizajes" + nl
        + "/profundo    Modo analisis profundo" + nl
        + "/normal      Modo respuestas rapidas" + nl
        + "/memoria     Lo que recuerdo de ti" + nl
        + "/clear       Borrar historial de conversacion" + nl + nl
        + "HABILIDADES (sin comando):" + nl
        + "URL de noticia        -> la leo + verifico si es fake" + nl
        + "'es esto verdad?'     -> Escudo de Veracidad" + nl
        + "'8 modos' / 'analiza' -> analisis profundo de contenido" + nl
        + "'busca todo sobre X'  -> KB + biblioteca simultaneo" + nl
        + "'que notas tengo de X'-> vault Obsidian" + nl
        + "'libros de [autor]'   -> biblioteca 2000+ libros" + nl
        + "'conecta esta nota'   -> grafo de conocimiento" + nl
        + "Reddit / HN           -> noticias tech en tiempo real" + nl
        + "Nota de voz           -> Whisper + respuesta por audio" + nl
        + "Foto                  -> vision Claude" + nl
        + "'estoy en Madrid'     -> clima de esa ciudad" + nl
        + "'agenda X para...'    -> Google Calendar/Tasks" + nl
    )
    await update.message.reply_text(capacidades, reply_markup=reply_markup)


async def button_handler(update, context):
    query = update.callback_query
    await query.answer()
    chat_id = update.effective_chat.id

    if query.data == 'btn_morning':
        await query.edit_message_text("☕ Preparando tu resumen matutino...")
        try:
            summary = await generate_morning_summary(chat_id)
            await send_long_message_raw(context, chat_id, summary)
        except Exception as e:
            await context.bot.send_message(chat_id, f"⚠️ Error: {e}")

    elif query.data == 'btn_news':
        await query.edit_message_text("📰 Preparando boletín de noticias...")
        try:
            from brain import generate_news_bulletin
            bulletin = await generate_news_bulletin()
            await send_long_message_raw(context, chat_id, bulletin)
        except Exception as e:
            await context.bot.send_message(chat_id, f"⚠️ Error: {e}")

    elif query.data == 'btn_deep':
        user_modes[chat_id] = "profundo"
        await query.edit_message_text("🧘‍♀️ Modo Profundo activado.")

    elif query.data == 'btn_normal':
        user_modes[chat_id] = "normal"
        await query.edit_message_text("⚡ Modo Normal activado.")

    elif query.data == 'btn_clear':
        conversation_history[chat_id] = []
        await query.edit_message_text("🧹 Chat reiniciado. (Memoria persistente intacta)")

    elif query.data == 'btn_mem':
        all_facts = get_all_facts() or {}
        lines = [f"â€¢ {k}: {v}" for k, v in all_facts.items() if not k.startswith("System_Location")]
        if lines:
            memory_text = "🧠 Lo que recuerdo de ti:\n\n" + "\n".join(lines)
            if len(memory_text) > 4000:
                memory_text = memory_text[:4000] + "\n\n[... Truncado]"
        else:
            memory_text = "🧠 Memoria vacía. Dime cosas y las recordaré."
        await query.edit_message_text(memory_text)

    elif query.data == 'btn_location':
        loc = user_locations.get(chat_id, DEFAULT_LOCATION)
        await query.edit_message_text(f"📍 {loc['name']} ({loc['lat']}, {loc['lng']})")

    elif query.data == 'btn_progreso':
        await query.edit_message_text("Generando panel de progreso...")
        try:
            from knowledge_base import kb_list, mental_models_stats
            from library import get_library_stats
            parts = [kb_list(mode='stats'), kb_list(mode='tags', limit=8), get_library_stats(), mental_models_stats(top_n=5)]
            await send_long_message_raw(context, chat_id, chr(10).join(parts))
        except Exception as e:
            await context.bot.send_message(chat_id, "Error: " + str(e))

    elif query.data == 'btn_sintesis':
        await query.edit_message_text("Generando sintesis semanal...")
        try:
            from brain import generate_weekly_synthesis
            synthesis = await generate_weekly_synthesis(chat_id)
            await send_long_message_raw(context, chat_id, synthesis)
        except Exception as e:
            await context.bot.send_message(chat_id, "Error: " + str(e))

    elif query.data == 'btn_img':
        await query.edit_message_text("🎨 Escríbeme qué imagen quieres que genere.")


# =====================================================
# =====================================================
# RECORDATORIOS PROACTIVOS
# =====================================================

async def check_reminders(context: ContextTypes.DEFAULT_TYPE):
    """Resumen matutino inteligente — UNA VEZ al día a las 9am CR. Pasa por Claude."""
    if not OWNER_CHAT_ID:
        return

    chat_id = int(OWNER_CHAT_ID)

    try:
        summary = await generate_morning_summary(chat_id)
        await send_long_message_raw(context, chat_id, summary)
        logger.info(f"☀️ Resumen matutino inteligente enviado a {chat_id}")
    except Exception as e:
        logger.error(f"Error resumen matutino: {e}")


# =====================================================
# HANDLERS PRINCIPALES
# =====================================================

@restricted
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler de mensajes de texto."""
    chat_id = update.effective_chat.id
    text = update.message.text

    # Feedback visual
    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    # Detectar solicitud de boletin matutino
    text_lower = text.lower()
    morning_triggers = ["resumen matutino", "boletin matutino", "buenos dias claudette",
                        "genera el boletin", "boletin del dia", "genera el resumen"]
    if any(t in text_lower for t in morning_triggers):
        await context.bot.send_chat_action(chat_id=chat_id, action="typing")
        summary = await generate_morning_summary(chat_id)
        await send_long_message(update, summary)
        return

    # Detectar mencion de ubicacion (G - Geolocalizacion inteligente)
    _loc_patterns = [
        r"(?:estoy en|llegue a|voy a|viajando a|desde) ([A-Za-z][a-zA-Z ]{2,30})(?:[,.]|$)",
    ]
    for _pat in _loc_patterns:
        _m = re.search(_pat, text, re.IGNORECASE)
        if _m:
            _city = _m.group(1).strip()
            _stopwords = ["que", "un", "el", "la", "mi", "tu", "se", "te", "me", "lo"]
            if not any(w == _city.lower() for w in _stopwords):
                user_locations[chat_id] = {"lat": 0, "lng": 0, "name": _city}
                try:
                    from tools_registry import get_weather_by_city as _gwc
                    _w = _gwc(_city)
                    text += chr(10) + chr(10) + "[SISTEMA: Pablo menciono estar en " + _city + ". Clima: " + _w + "]"
                except Exception:
                    text += chr(10) + chr(10) + "[SISTEMA: Ubicacion actualizada a " + _city + ".]"
            break

    # Detectar YouTube
    yt_transcript = get_youtube_transcript(text)
    if yt_transcript:
        text += f"\n\n[SISTEMA]: El usuario envió un video. {yt_transcript}"

    # Procesar con brain
    response = await process_chat(update, context, text)

    # Enviar respuesta (split si es muy larga para Telegram)
    await send_long_message(update, response)


@restricted
async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler de notas de voz → Whisper → Claude → ElevenLabs."""
    if not openai_client:
        return await update.message.reply_text("Whisper no configurado.")
    try:
        file = await context.bot.get_file(update.message.voice.file_id)
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as f:
            await file.download_to_drive(f.name)
            path = f.name

        with open(path, "rb") as audio_file:
            transcript = openai_client.audio.transcriptions.create(
                model="whisper-1", file=audio_file
            ).text
        os.unlink(path)

        await update.message.reply_text(f"🎤 {transcript}")
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

        response = await process_chat(update, context, transcript)
        await send_long_message(update, response)

        # Respuesta por voz
        if elevenlabs_client:
            try:
                text_clean = re.sub(r'[^\w\s,.?¡!]', '', response)
                audio = elevenlabs_client.text_to_speech.convert(
                    text=text_clean,
                    voice_id=ELEVENLABS_VOICE_ID,
                    model_id="eleven_multilingual_v2"
                )
                await update.effective_message.reply_voice(voice=b"".join(audio))
            except Exception as e:
                logger.error(f"âŒ ElevenLabs Error: {e}")

    except Exception as e:
        await update.message.reply_text(f"Error voz: {e}")


@restricted
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler de fotos → visión Claude."""
    photo_file = await update.message.photo[-1].get_file()
    with io.BytesIO() as f:
        await photo_file.download_to_memory(out=f)
        image_data = base64.b64encode(f.getvalue()).decode("utf-8")

    caption = update.message.caption or "Analiza esta imagen."

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    response = await process_chat(update, context, caption, image_data=image_data)
    await send_long_message(update, response)


@restricted
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler de documentos — extrae contenido de archivos enviados por Telegram."""
    doc = update.message.document
    if not doc:
        return

    chat_id = update.effective_chat.id
    file_name = doc.file_name or "archivo"
    file_size = doc.file_size or 0
    caption = update.message.caption or ""

    # Límite de tamaño: 20MB (Telegram permite hasta 20MB para bots)
    if file_size > 20 * 1024 * 1024:
        await update.message.reply_text("⚠️ Archivo demasiado grande (máx 20MB).")
        return

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    # Determinar tipo y extensión
    lower_name = file_name.lower()
    supported_extensions = ('.txt', '.md', '.csv', '.json', '.py', '.js', '.html',
                            '.pdf', '.docx', '.xlsx', '.xls')

    if not any(lower_name.endswith(ext) for ext in supported_extensions):
        # Tipo no soportado — pasar solo el nombre
        msg_text = f"El usuario envió un archivo: {file_name}"
        if caption:
            msg_text += f"\nMensaje: {caption}"
        msg_text += "\n(Formato no soportado para lectura directa. Sugerí subirlo a Drive.)"
        response = await process_chat(update, context, msg_text)
        await send_long_message(update, response)
        return

    try:
        # Descargar archivo
        tg_file = await context.bot.get_file(doc.file_id)
        file_bytes = io.BytesIO()
        await tg_file.download_to_memory(out=file_bytes)
        file_bytes.seek(0)

        extracted_text = ""

        # --- TEXT-BASED: .txt, .md, .csv, .json, .py, .js, .html ---
        if any(lower_name.endswith(ext) for ext in ('.txt', '.md', '.csv', '.json', '.py', '.js', '.html')):
            try:
                extracted_text = file_bytes.read().decode('utf-8', errors='ignore')
            except Exception:
                extracted_text = file_bytes.read().decode('latin-1', errors='ignore')

        # --- PDF ---
        elif lower_name.endswith('.pdf'):
            try:
                import pypdf
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                    tmp.write(file_bytes.read())
                    tmp_path = tmp.name
                reader = pypdf.PdfReader(tmp_path)
                pages = min(len(reader.pages), 50)
                extracted_text = "\n".join(
                    reader.pages[i].extract_text() or "" for i in range(pages)
                )
                os.unlink(tmp_path)
            except Exception as e:
                extracted_text = f"(Error leyendo PDF: {e})"

        # --- DOCX ---
        elif lower_name.endswith('.docx'):
            try:
                from docx import Document as DocxDoc
                with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as tmp:
                    tmp.write(file_bytes.read())
                    tmp_path = tmp.name
                doc_obj = DocxDoc(tmp_path)
                extracted_text = "\n".join(p.text for p in doc_obj.paragraphs if p.text.strip())
                os.unlink(tmp_path)
            except Exception as e:
                extracted_text = f"(Error leyendo DOCX: {e})"

        # --- XLSX / XLS ---
        elif lower_name.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(file_bytes.read())
                    tmp_path = tmp.name
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                sheets_text = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        row_str = " | ".join(str(c) if c is not None else "" for c in row)
                        if row_str.strip(" |"):
                            rows.append(row_str)
                    if rows:
                        sheets_text.append(f"--- Hoja: {sheet_name} ---\n" + "\n".join(rows))
                extracted_text = "\n\n".join(sheets_text)
                wb.close()
                os.unlink(tmp_path)
            except Exception as e:
                extracted_text = f"(Error leyendo Excel: {e})"

        # Truncar si es muy largo
        if len(extracted_text) > 12000:
            extracted_text = extracted_text[:12000] + "\n\n[... Contenido truncado. Pedí una sección específica.]"

        # Construir mensaje para Claude con el contenido del archivo
        msg_text = f"🔎 El usuario envió el archivo '{file_name}'.\n"
        if caption:
            msg_text += f"Mensaje: {caption}\n"
        msg_text += f"\n--- CONTENIDO DEL ARCHIVO ---\n{extracted_text}\n--- FIN DEL ARCHIVO ---"

        response = await process_chat(update, context, msg_text)
        await send_long_message(update, response)

    except Exception as e:
        logger.error(f"Document handler error: {e}", exc_info=True)
        await update.message.reply_text(f"⚠️ Error procesando archivo: {e}")


async def handle_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler de ubicación compartida."""
    msg = update.effective_message
    if not msg or not msg.location:
        return

    lat, lng = msg.location.latitude, msg.location.longitude
    chat_id = update.effective_chat.id

    user_locations[chat_id] = {"lat": lat, "lng": lng, "name": "Ubicación Telegram"}

    try:
        save_fact(f"System_Location_Lat_{chat_id}", str(lat))
        save_fact(f"System_Location_Lng_{chat_id}", str(lng))
        logger.info(f"💾 Ubicación guardada: {lat}, {lng}")
    except Exception as e:
        logger.error(f"Error guardando ubicación: {e}")

    if not update.edited_message:
        await msg.reply_text("📍 Ubicación actualizada.")


# =====================================================
# COMANDOS DIRECTOS
# =====================================================

@restricted
async def cmd_buenos_dias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    await update.message.reply_text("☕ Preparando tu resumen matutino...")
    try:
        summary = await generate_morning_summary(chat_id)
        await send_long_message(update, summary)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Error: {e}")


@restricted
async def cmd_noticias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📰 Preparando boletín de noticias...")
    try:
        from brain import generate_news_bulletin
        bulletin = await generate_news_bulletin()
        await send_long_message(update, bulletin)
    except Exception as e:
        await update.message.reply_text(f"⚠️ Error: {e}")


@restricted
async def cmd_profundo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_modes[update.effective_chat.id] = "profundo"
    await update.message.reply_text("🧘‍♀️ Modo Profundo activado.")


@restricted
async def cmd_normal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_modes[update.effective_chat.id] = "normal"
    await update.message.reply_text("⚡ Modo Normal activado.")


@restricted
async def cmd_clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conversation_history[update.effective_chat.id] = []
    await update.message.reply_text("🧹 Memoria de conversación borrada. (Memoria persistente intacta)")


@restricted
async def cmd_memoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    all_facts = get_all_facts() or {}
    lines = [f"â€¢ {k}: {v}" for k, v in all_facts.items() if not k.startswith("System_Location")]
    if lines:
        memory_text = "🧠 Lo que recuerdo de ti:\n\n" + "\n".join(lines)
        if len(memory_text) > 4000:
            memory_text = memory_text[:4000] + "\n\n[... Truncado]"
    else:
        memory_text = "🧠 Memoria vacía. Dime cosas y las recordaré."
    await update.message.reply_text(memory_text)


@restricted
async def cmd_progreso(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Panel de progreso: KB stats + Biblioteca + insights guardados."""
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    lines = ["=== PANEL DE PROGRESO - CLAUDETTE ===\n"]

    # KB stats (Vault Obsidian)
    try:
        from knowledge_base import kb_list
        lines.append(kb_list(mode="stats"))
    except Exception as e:
        lines.append(f"KB: error ({e})")

    # Tags mas usados (top 10)
    try:
        from knowledge_base import kb_list
        lines.append("\n" + kb_list(mode="tags", limit=10))
    except Exception:
        pass

    # Biblioteca stats
    try:
        from library import get_library_stats
        lines.append("\n" + get_library_stats())
    except Exception as e:
        lines.append(f"\nBiblioteca: error ({e})")

    # Memoria rapida (hechos guardados)
    try:
        all_facts = get_all_facts() or {}
        fact_count = len([k for k in all_facts if not k.startswith("System_")])
        lines.append(f"\nMemoria rapida: {fact_count} hechos guardados")
    except Exception:
        pass

    # Ultimas notas indexadas
    try:
        from knowledge_base import kb_list
        lines.append("\n" + kb_list(mode="recent", limit=5))
    except Exception:
        pass

    await send_long_message(update, "\n".join(lines))


@restricted
async def cmd_sintesis(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Genera la sintesis semanal bajo demanda."""
    await update.message.reply_text("Generando sintesis semanal...")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    try:
        synthesis = await generate_weekly_synthesis(update.effective_chat.id)
        await send_long_message(update, synthesis)
    except Exception as e:
        await update.message.reply_text(f"Error: {e}")


# =====================================================
# UTILIDADES
# =====================================================

async def send_long_message(update, text, max_length=4000):
    """Envía mensajes largos dividiéndolos si superan el límite de Telegram."""
    if len(text) <= max_length:
        await update.effective_message.reply_text(text)
        return

    # Dividir en chunks respetando saltos de línea
    chunks = []
    while text:
        if len(text) <= max_length:
            chunks.append(text)
            break
        # Buscar un buen punto de corte
        cut = text.rfind('\n', 0, max_length)
        if cut == -1:
            cut = max_length
        chunks.append(text[:cut])
        text = text[cut:].lstrip('\n')

    for chunk in chunks:
        if chunk.strip():
            await update.effective_message.reply_text(chunk)


async def send_long_message_raw(context, chat_id, text, max_length=4000):
    """Envía mensajes largos usando context.bot directamente (sin Update)."""
    if len(text) <= max_length:
        await context.bot.send_message(chat_id=chat_id, text=text)
        return

    chunks = []
    while text:
        if len(text) <= max_length:
            chunks.append(text)
            break
        cut = text.rfind('\n', 0, max_length)
        if cut == -1:
            cut = max_length
        chunks.append(text[:cut])
        text = text[cut:].lstrip('\n')

    for chunk in chunks:
        if chunk.strip():
            await context.bot.send_message(chat_id=chat_id, text=chunk)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error(msg="Exception:", exc_info=context.error)


# =====================================================
# AUTO-LOG DE DESARROLLO
# =====================================================

def _log_dev_commits_to_kb():
    """
    Lee los ultimos 30 commits de git y guarda los nuevos en la KB bajo
    la categoria 'claudette_dev'. Usa el SHA como deduplicacion para no
    repetir entradas ya guardadas.
    """
    import subprocess
    from knowledge_base import kb_save_insight, _get_conn

    vault_path = os.environ.get("OBSIDIAN_VAULT_PATH")
    if not vault_path:
        return

    # Leer commits recientes
    result = subprocess.run(
        ["git", "log", "--oneline", "-30", "--no-merges"],
        capture_output=True, text=True,
        cwd=os.path.dirname(os.path.abspath(__file__))
    )
    if result.returncode != 0 or not result.stdout.strip():
        return

    lines = result.stdout.strip().splitlines()

    # Verificar cuales SHAs ya existen en KB para no duplicar
    try:
        conn = _get_conn()
        cur = conn.cursor()
        cur.execute("SELECT content FROM knowledge_chunks WHERE source ILIKE '%CLAUDETTE_MEMORY%'")
        existing_text = " ".join(row[0] for row in cur.fetchall() if row[0])
        conn.close()
    except Exception:
        existing_text = ""

    saved = 0
    for line in reversed(lines):  # de mas antiguo a mas nuevo
        if not line.strip():
            continue
        parts = line.split(" ", 1)
        if len(parts) < 2:
            continue
        sha, msg = parts[0], parts[1]

        # Deduplicar por SHA
        if sha in existing_text:
            continue

        kb_save_insight(
            category="claudette_dev",
            title=msg,
            content=f"Commit: `{sha}`\nCambio: {msg}",
            project="Claudette"
        )
        saved += 1

    if saved:
        logger.info(f"Auto-log: {saved} commits guardados en KB (claudette_dev)")


# =====================================================
# MAIN
# =====================================================

def main():
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    # Comandos
    app.add_handler(CommandHandler("start", show_menu))
    app.add_handler(CommandHandler("menu", show_menu))
    app.add_handler(CommandHandler("clear", cmd_clear))
    app.add_handler(CommandHandler("profundo", cmd_profundo))
    app.add_handler(CommandHandler("normal", cmd_normal))
    app.add_handler(CommandHandler("buenosdias", cmd_buenos_dias))
    app.add_handler(CommandHandler("noticias", cmd_noticias))
    app.add_handler(CommandHandler("memoria", cmd_memoria))
    app.add_handler(CommandHandler("progreso", cmd_progreso))
    app.add_handler(CommandHandler("sintesis", cmd_sintesis))

    # Botones inline
    app.add_handler(CallbackQueryHandler(button_handler))

    # Mensajes
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.LOCATION, handle_location))

    # Recordatorio matutino — UNA SOLA VEZ al día a las 6:00 AM Costa Rica
    try:
        if OWNER_CHAT_ID and app.job_queue:
            from datetime import time as dt_time
            # 6:00 AM Costa Rica = 12:00 UTC (GMT-6)
            morning_time = dt_time(hour=12, minute=0, second=0)
            app.job_queue.run_daily(
                check_reminders,
                time=morning_time,
                name="morning_reminder"
            )
            logger.info(f"🔔 Recordatorio matutino (6:00 AM CR) activado para chat_id: {OWNER_CHAT_ID}")
        else:
            logger.warning("⚠️ Recordatorios desactivados (falta OWNER_CHAT_ID o job-queue).")
    except Exception as e:
        logger.warning(f"⚠️ Recordatorios no disponibles: {e}")

    # Sintesis semanal - domingos 6pm Costa Rica = lunes 00:00 UTC
    try:
        if OWNER_CHAT_ID and app.job_queue:
            from datetime import time as dt_time

            async def send_weekly_synthesis(context):
                chat_id = int(OWNER_CHAT_ID)
                try:
                    synthesis = await generate_weekly_synthesis(chat_id)
                    await send_long_message_raw(context, chat_id, synthesis)
                    logger.info("Sintesis semanal enviada")
                except Exception as e:
                    logger.error(f"Sintesis semanal error: {e}")

            app.job_queue.run_daily(
                send_weekly_synthesis,
                time=dt_time(hour=0, minute=0, second=0),
                days=(0,),  # 0 = lunes (lunes 00:00 UTC = domingo 6pm CR)
                name="weekly_synthesis"
            )
    except Exception as e:
        logger.warning(f"Sintesis semanal no disponible: {e}")

    # Memoria proactiva — cada 3 días a las 7pm Costa Rica (01:00 UTC)
    try:
        if OWNER_CHAT_ID and app.job_queue:
            from datetime import time as dt_time
            from brain import check_patterns_proactive

            app.job_queue.run_daily(
                check_patterns_proactive,
                time=dt_time(hour=1, minute=0, second=0),
                days=(1, 4),  # martes y viernes 01:00 UTC = lunes/jueves 7pm CR
                name="proactive_memory"
            )
            logger.info("🔍 Memoria proactiva activada (martes y viernes 7pm CR)")
    except Exception as e:
        logger.warning(f"Memoria proactiva no disponible: {e}")

    # Registrar comandos en Telegram (aparecen al escribir /)
    try:
        commands = [
            BotCommand("buenosdias",  "Boletin matutino: noticias + Midas + libro del dia"),
            BotCommand("noticias",    "Noticias curadas en tiempo real (RSS + HN)"),
            BotCommand("progreso",    "Panel: KB Obsidian + biblioteca + modelos mentales"),
            BotCommand("sintesis",    "Sintesis semanal de aprendizajes e insights"),
            BotCommand("profundo",    "Activar modo analisis profundo"),
            BotCommand("normal",      "Volver al modo respuestas rapidas"),
            BotCommand("memoria",     "Ver datos que Claudette recuerda de ti"),
            BotCommand("clear",       "Borrar historial de conversacion"),
            BotCommand("menu",        "Menu completo con todas las habilidades"),
            BotCommand("start",       "Menu completo con todas las habilidades"),
        ]
        import asyncio
        async def _set_cmds():
            await app.bot.set_my_commands(commands)
        asyncio.get_event_loop().run_until_complete(_set_cmds())
        logger.info("Comandos Telegram registrados OK")
    except Exception as e:
        logger.warning("set_my_commands error: " + str(e))

    app.add_error_handler(error_handler)

    # Auto-log de commits de desarrollo al vault
    try:
        _log_dev_commits_to_kb()
    except Exception as e:
        logger.warning(f"Auto-log dev commits: {e}")

    print("🚀 Claudette 2.0 (Modular + YouTube + Google Services) ONLINE")
    app.run_polling()


if __name__ == '__main__':
    main()

